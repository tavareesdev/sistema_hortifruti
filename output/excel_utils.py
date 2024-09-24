import pandas as pd
import sys
import openpyxl
import time
import re
from datetime import datetime

# Caminho para o arquivo da planilha
planilha_caminho = 'C:\\Users\\gtava\\OneDrive\\Documentos\\project\\output\\database.xlsx'

"""
Lê dados de uma aba específica de uma planilha Excel e retorna um DataFrame.

Parâmetros:
- sheet_name (str): Nome da aba da planilha a ser lida.

Retorna:
- pandas.DataFrame: Dados da aba, se a leitura for bem-sucedida.
- None: Se ocorrer um erro, como arquivo não encontrado.

Exceções:
- FileNotFoundError: Arquivo não encontrado.
- Exception: Outros erros ao ler o arquivo.
"""
def ler_dados_planilha(sheet_name):
    try:
        df = pd.read_excel(planilha_caminho, sheet_name=sheet_name, dtype=str)
        return df
    except FileNotFoundError:
        print(f"Arquivo não encontrado: {planilha_caminho}")
        return None
    except Exception as e:
        print(f"Erro ao ler o arquivo: {e}")
        return None

"""
Lê os dados de uma aba de uma planilha Excel.

Parâmetros:
- sheet_name (str): Nome da aba a ser lida.

Retorna:
- pandas.DataFrame: Dados da aba se a leitura for bem-sucedida.
- None: Se ocorrer um erro, como arquivo não encontrado ou outro problema.

Exceções:
- FileNotFoundError: Arquivo não encontrado.
- Exception: Qualquer outro erro ao ler o arquivo.
"""
def escrever_dados_planilha(df, sheet_name):
    try:
        with pd.ExcelWriter(planilha_caminho, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    except Exception as e:
        print(f"Erro ao escrever no arquivo: {e}")

"""
Gera um novo ID único para adicionar a um DataFrame.

Parâmetros:
- df (pandas.DataFrame): DataFrame que contém uma coluna de IDs.

Retorna:
- int: Novo ID único. Começa em 1 se o DataFrame estiver vazio ou se a coluna 'ID' não existir.

Comportamento:
- Se o DataFrame estiver vazio ou não tiver uma coluna 'ID', retorna 1.
- Se a coluna 'ID' não contiver IDs válidos, também retorna 1.
- Caso contrário, retorna o maior ID existente incrementado de 1.
"""
def gerar_id_unico(df):
    """Gera um novo ID único."""
    if df.empty or 'ID' not in df.columns:
        return 1
    else:
        ids_existentes = pd.to_numeric(df['ID'], errors='coerce').dropna().astype(int)
        if ids_existentes.empty:
            return 1
        else:
            return ids_existentes.max() + 1

"""
Cadastra um novo funcionário e usuário nas planilhas.

Parâmetros:
- nome (str): Nome do funcionário.
- cpf (str): CPF do funcionário.
- data_nascimento (str): Data de nascimento do funcionário.
- rg (str): RG do funcionário.
- username (str): Nome de usuário para login.
- password (str): Senha para login.
- tipo (str): Tipo de usuário (e.g., 'admin', 'funcionario').

Comportamento:
- Lê os dados das planilhas 'Funcionarios' e 'Users'.
- Verifica se o CPF ou RG já existem na planilha 'Funcionarios' ou se o Username já existe na planilha 'Users'.
- Se algum desses dados já existir, exibe uma mensagem de erro e aguarda 3 segundos.
- Caso contrário, gera um novo ID único, adiciona o funcionário e o usuário às respectivas planilhas e grava as atualizações de volta.
"""
def cadastrar_funcionario(nome, cpf, data_nascimento, rg, username, password, tipo):
    # Lê os dados atuais das planilhas
    df_funcionarios = ler_dados_planilha('Funcionarios')
    df_users = ler_dados_planilha('Users')

    if df_funcionarios is not None and df_users is not None:
        # Verifica se o CPF ou RG já existem na planilha 'Funcionarios'
        cpf_existente = df_funcionarios['CPF'].astype(str).str.contains(str(cpf)).any()
        rg_existente = df_funcionarios['RG'].astype(str).str.contains(str(rg)).any()

        # Verifica se o Username já existe na planilha 'Users'
        username_existente = df_users['Username'].astype(str).str.contains(str(username)).any()

        if cpf_existente:
            print("Erro: Já existe um funcionario cadastrado com este CPF. Voltando a tela de cadastro...")
            time.sleep(3)
        elif rg_existente:
            print("Erro: Já existe um funcionario cadastrado com este RG. Voltando a tela de cadastro...")
            time.sleep(3)
        elif username_existente:
            print("Erro: Já existe um usuario cadastrado com este Username. Voltando a tela de cadastro...")
            time.sleep(3)
        else:
            # Gera um novo ID único
            novo_id = gerar_id_unico(df_funcionarios)

            # Adiciona o novo funcionário à planilha 'Funcionarios'
            novo_funcionario = pd.DataFrame({
                'ID': [novo_id],
                'Nome': [nome],
                'CPF': [cpf],
                'Data de Nascimento': [data_nascimento],
                'RG': [rg]
            })
            df_funcionarios = pd.concat([df_funcionarios, novo_funcionario], ignore_index=True)

            # Adiciona o novo usuário à planilha 'Users'
            novo_usuario = pd.DataFrame({
                'ID': [novo_id],
                'Username': [username],
                'Password': [password],
                'Tipo': [tipo]
            })
            df_users = pd.concat([df_users, novo_usuario], ignore_index=True)

            # Escreve os dados de volta nas planilhas
            escrever_dados_planilha(df_funcionarios, 'Funcionarios')
            escrever_dados_planilha(df_users, 'Users')

            print("Cadastro realizado com sucesso!")
    else:
        print("Erro ao ler as planilhas.")

"""
Cadastra um novo cliente na planilha 'Clientes'.

Parâmetros:
- nome (str): Nome do cliente.
- cpf (str): CPF do cliente.
- data_nascimento (str): Data de nascimento do cliente.
- rg (str): RG do cliente.

Comportamento:
- Lê os dados da planilha 'Clientes'.
- Verifica se o CPF ou RG já existem na planilha.
- Se CPF ou RG já estiverem cadastrados, exibe uma mensagem de erro e aguarda 3 segundos.
- Caso contrário, gera um novo ID único, adiciona o cliente à planilha e grava as alterações.
"""
def cadastrar_cliente(nome, cpf, data_nascimento, rg):
    # Lê os dados atuais das planilhas
    df_cliente = ler_dados_planilha('Clientes')

    if df_cliente is not None:
         # Verifica se o CPF ou RG já existem na planilha 'Clientes'
        cpf_existente = df_cliente['CPF'].astype(str).str.contains(str(cpf)).any()
        rg_existente = df_cliente['RG'].astype(str).str.contains(str(rg)).any()

        if cpf_existente:
            print("Erro: Já existe um cliente cadastrado com este CPF. Voltando a tela de cadastro...")
            time.sleep(3)
        elif rg_existente:
            print("Erro: Já existe um cliente cadastrado com este RG. Voltando a tela de cadastro...")
            time.sleep(3)
        else:
            novo_id = gerar_id_unico(df_cliente)
            # Adiciona o novo funcionário à planilha 'cliente'
            novo_cliente = pd.DataFrame({
                'ID': [novo_id],
                'Nome': [nome],
                'CPF': [cpf],
                'Data de Nascimento': [data_nascimento],
                'RG': [rg]
            })
            df_cliente = pd.concat([df_cliente, novo_cliente], ignore_index=True)

            # Escreve os dados de volta nas planilhas
            escrever_dados_planilha(df_cliente, 'Clientes')

            print("Cadastro realizado com sucesso!")
    else:
        print("Erro ao ler as planilhas.")

"""
Cadastra um novo produto na planilha 'Produtos'.

Parâmetros:
- nome (str): Nome do produto.
- preco (str): Preço do produto, deve ser um número válido.
- qtd_produto (int): Quantidade do produto em estoque.
- tipo_produto (str): Tipo de venda do produto (ex: unidade, kg).

Comportamento:
- Lê os dados da planilha 'Produtos'.
- Gera um novo ID único para o produto.
- Converte o preço para float e o arredonda para duas casas decimais.
- Adiciona o novo produto à planilha e grava as alterações.
- Se o preço não for válido, exibe uma mensagem de erro.
"""
def cadastrar_produto(nome, preco, qtd_produto, tipo_produto):
    # Lê os dados atuais das planilhas
    df_produto = ler_dados_planilha('Produtos')

    # Gera um novo ID único
    novo_id = gerar_id_unico(df_produto)

    if df_produto is not None:
        try:
            # Converte o preço para float e arredonda para duas casas decimais
            preco_float = float(preco)  # Converte para float
            preco_arredondado = round(preco_float, 2)  # Arredonda para duas casas decimais
        except ValueError:
            print("Erro: O preço deve ser um número válido.")
            return
        
        # Adiciona o novo produto à planilha 'produto'
        novo_produto = pd.DataFrame({
            'ID': [novo_id],
            'Nome': [nome],
            'Preco': [preco_arredondado],
            'Quantidade': [qtd_produto],
            'Tipo de Venda': [tipo_produto]
        })
        df_produto = pd.concat([df_produto, novo_produto], ignore_index=True)

        # Escreve os dados de volta nas planilhas
        escrever_dados_planilha(df_produto, 'Produtos')

        print("Cadastro realizado com sucesso!")
    else:
        print("Erro ao ler as planilhas.")

"""
Valida o login de um usuário verificando as credenciais na planilha 'Users'.

Parâmetros:
- username (str): Nome de usuário para login.
- password (str): Senha do usuário para login.

Retorno:
- bool: Retorna True se o login for bem-sucedido, caso contrário, retorna False.

Comportamento:
- Lê os dados da planilha 'Users'.
- Verifica se existe um usuário com o nome de usuário e senha fornecidos.
- Exibe uma mensagem de sucesso se as credenciais forem corretas, ou uma mensagem de erro se forem incorretas.
- Se houver um erro ao ler a planilha, exibe uma mensagem de erro e retorna False.
"""
def validar_login(username, password):
    df_users = ler_dados_planilha('Users')
    if df_users is not None:
        user_exists = df_users[(df_users['Username'] == username) & (df_users['Password'] == password)]
        if not user_exists.empty:
            user_info = user_exists.iloc[0]  # Pega a primeira linha correspondente
            df_users2 = ler_dados_planilha('Funcionarios')
            user_exists2 = df_users2[(df_users2['ID'] == user_info['ID'])]
            user_info2 = user_exists2.iloc[0]  # Pega a primeira linha correspondente
            print(f"{user_info['ID']},{user_info2['Nome']},{user_info['Tipo']}")
            return True
        else:
            print("Usuario ou senha incorretos.")
            return False
    else:
        print("Erro ao ler a planilha de usuários.")
        return False

"""
Busca e exibe produtos na planilha 'Produtos' com base no nome fornecido.

Parâmetros:
- nome (str): Nome do produto para a busca. Se for 'Todos', retorna todos os produtos.

Comportamento:
- Lê os dados da planilha 'Produtos'.
- Se o nome não for 'Todos', busca produtos cujo nome contenha o texto fornecido, ignorando maiúsculas e minúsculas.
- Se o nome for 'Todos', exibe todos os produtos.
- Exibe os resultados encontrados ou uma mensagem informando que nenhum produto foi encontrado.
- Se houver um erro ao carregar os dados ou se o nome fornecido não for uma string válida, exibe uma mensagem de erro.
"""
def buscar_produto(nome):
    df = ler_dados_planilha('Produtos')
    
    if df is None:
        print("Erro ao carregar os dados da planilha.")
        return

    if nome != 'Todos':
        if not isinstance(nome, str):
            print("O nome do produto deve ser uma string.")
            return
        resultado = df[df['Nome'].str.contains(nome.strip(), case=False, na=False)]
    else:
        resultado = df

    if resultado is not None and not resultado.empty:
        resultado_str = resultado.to_string(index=False)
        print("\n" + resultado_str + "\n")
    else:
        print("Nenhum produto encontrado.")

"""
Atualiza os detalhes de um produto na planilha 'Produtos' com base no ID fornecido.

Parâmetros:
- novo_nome (str): Novo nome do produto. Se não fornecido, o nome atual não é alterado.
- novo_preco (float): Novo preço do produto. Se não fornecido, o preço atual não é alterado.
- nova_quantidade (int): Nova quantidade do produto. Se não fornecida, a quantidade atual não é alterada.
- novo_tipo (str): Novo tipo de venda do produto. Se não fornecido, o tipo atual não é alterado.
- id_produto (int): ID do produto a ser atualizado.

Comportamento:
- Lê os dados da planilha 'Produtos'.
- Localiza o produto pelo ID fornecido.
- Atualiza os campos do produto com os novos valores fornecidos, se forem informados.
- Salva as alterações de volta na planilha 'Produtos'.
- Exibe uma mensagem de sucesso se o produto for atualizado com sucesso.
- Exibe uma mensagem de erro se o produto com o ID especificado não for encontrado ou se houver um erro ao ler a planilha.
"""
def editar_produto(novo_nome, novo_preco, nova_quantidade, novo_tipo, id_produto):
    # Lê os dados atuais da planilha 'Produtos'
    df_produto = ler_dados_planilha('Produtos')
    print(id_produto)
    if df_produto is not None:
        # Localiza o produto pelo ID
        produto_idx = df_produto[df_produto['ID'] == id_produto].index

        if not produto_idx.empty:
            # Atualiza os campos se o usuário forneceu novos valores
            if novo_nome:
                df_produto.at[produto_idx[0], 'Nome'] = novo_nome
            if novo_preco:
                df_produto.at[produto_idx[0], 'Preco'] = round(float(novo_preco), 2)
            if nova_quantidade:
                df_produto.at[produto_idx[0], 'Quantidade'] = int(nova_quantidade)
            if novo_tipo:
                df_produto.at[produto_idx[0], 'Tipo de Venda'] = novo_tipo

            # Salva as alterações na planilha
            escrever_dados_planilha(df_produto, 'Produtos')

            print("Produto atualizado com sucesso!")
        else:
            print("Produto com o ID especificado não foi encontrado.")
    else:
        print("Erro ao ler a planilha.")

"""
Exclui um produto da planilha 'Produtos' com base no ID fornecido.

Parâmetros:
- id_produto (int): ID do produto a ser excluído.

Comportamento:
- Carrega a planilha Excel especificada em `planilha_caminho`.
- Seleciona a planilha 'Produtos'.
- Percorre as linhas da planilha para encontrar a linha com o ID especificado.
- Exclui a linha correspondente ao ID fornecido.
- Salva as alterações no arquivo Excel.
- Exibe uma mensagem de sucesso após a exclusão do produto.
"""
def excluir_produto(id_produto):
    # Carregar o arquivo Excel
    wb = openpyxl.load_workbook(planilha_caminho)
    
    # Selecionar a planilha "produtos"
    ws = wb['Produtos']
    
    # Percorrer as linhas da planilha a partir da segunda linha
    for row in ws.iter_rows(min_row=2):
        # Acessar o valor na primeira coluna (ID) e garantir que ambos são strings para comparação
        id_valor = str(row[0].value)
        id_produto_str = str(id_produto)
        
        # Verificar se o ID corresponde ao que queremos excluir
        if id_valor == id_produto_str:
            # Excluir a linha
            ws.delete_rows(row[0].row)
            break
    
    # Salvar o arquivo sem afetar outras planilhas
    wb.save(planilha_caminho)
    print(f"Produto foi excluido com sucesso.")

"""
Busca funcionários na planilha 'Funcionarios' pelo nome fornecido.

Parâmetros:
- nome (str): Nome do funcionário a ser pesquisado. Se for 'Todos', retorna todos os funcionários.

Comportamento:
- Lê os dados da planilha 'Funcionarios'.
- Se `nome` não for 'Todos', verifica se `nome` é uma string e realiza uma busca parcial por nome.
- Se `nome` for 'Todos', retorna todos os funcionários.
- Exibe os resultados da busca ou uma mensagem informando que nenhum funcionário foi encontrado.
"""
def buscar_funcionario(nome):
    df = ler_dados_planilha('Funcionarios')

    if nome != 'Todos':
        if not isinstance(nome, str):
            print("O nome do funcionario deve ser uma string.")
            return
        resultado = df[df['Nome'].str.contains(nome.strip(), case=False, na=False)]
    else:
        resultado = df

    if not resultado.empty:
        resultado = resultado.to_string(index=False)
        print("\n" + resultado + "\n")
    else:
        print("Nenhum funcionario encontrado.")

"""
Atualiza as informações de um funcionário nas planilhas 'Funcionarios' e 'Users'.

Parâmetros:
- novo_nome (str): Novo nome do funcionário (opcional).
- novo_cpf (str): Novo CPF do funcionário (opcional).
- novo_rg (str): Novo RG do funcionário (opcional).
- nova_data_nasc (str): Nova data de nascimento do funcionário (opcional).
- novo_user (str): Novo nome de usuário do funcionário (opcional).
- nova_senha (str): Nova senha do funcionário (opcional).
- id_funcionario (int): ID do funcionário a ser atualizado.

Comportamento:
- Lê os dados das planilhas 'Funcionarios' e 'Users'.
- Atualiza os dados do funcionário nas planilhas se o ID correspondente for encontrado e os novos valores forem fornecidos.
- Salva as alterações nas planilhas.
- Exibe uma mensagem de sucesso se as atualizações forem bem-sucedidas, ou mensagens de erro caso contrário.
"""
def editar_funcionario(novo_nome, novo_cpf, novo_rg, nova_data_nasc, novo_user, nova_senha, id_funcionario):
    # Lê os dados atuais da planilha 'funcionarios'
    df_funcionario = ler_dados_planilha('Funcionarios')

    if df_funcionario is not None:
        # Localiza o funcionario pelo ID
        funcionario_idx = df_funcionario[df_funcionario['ID'] == id_funcionario].index

        if not funcionario_idx.empty:
            # Atualiza os campos se o usuário forneceu novos valores
            if novo_nome:
                df_funcionario.at[funcionario_idx[0], 'Nome'] = novo_nome
            if novo_cpf:
                df_funcionario.at[funcionario_idx[0], 'CPF'] = novo_cpf
            if novo_rg:
                df_funcionario.at[funcionario_idx[0], 'RG'] = novo_rg
            if nova_data_nasc:
                df_funcionario.at[funcionario_idx[0], 'Data de Nascimento'] = nova_data_nasc

            # Salva as alterações na planilha
            escrever_dados_planilha(df_funcionario, 'Funcionarios')

            a =1
        else:
            print("Funcionario com o ID especificado não foi encontrado.")
    else:
        print("Erro ao ler a planilha.")

    df_funcionario = ler_dados_planilha('Users')

    if df_funcionario is not None:
        # Localiza o funcionario pelo ID
        funcionario_idx = df_funcionario[df_funcionario['ID'] == id_funcionario].index

        if not funcionario_idx.empty:
            # Atualiza os campos se o usuário forneceu novos valores
            if novo_user:
                df_funcionario.at[funcionario_idx[0], 'Username'] = novo_user
            if nova_senha:
                df_funcionario.at[funcionario_idx[0], 'Password'] = nova_senha

            # Salva as alterações na planilha
            escrever_dados_planilha(df_funcionario, 'Users')

            a =1 
        else:
            print("funcionario com o ID especificado não foi encontrado.")
    else:
        print("Erro ao ler a planilha.")

    if a == 1:
        print("Funcionario atualizado com sucesso!")

"""
Exclui um funcionário das planilhas 'Funcionarios' e 'Users' com base no ID fornecido.

Parâmetros:
- id_funcionario (int): ID do funcionário a ser excluído.

Comportamento:
- Carrega o arquivo Excel e seleciona a planilha 'Funcionarios'.
- Percorre as linhas da planilha para encontrar e excluir a linha com o ID correspondente.
- Salva as alterações na planilha 'Funcionarios'.
- Recarrega o arquivo Excel, seleciona a planilha 'Users', e repete o processo de exclusão para garantir que o funcionário seja removido de ambas as planilhas.
- Salva as alterações na planilha 'Users'.
- Exibe uma mensagem de sucesso após a exclusão.
"""
def excluir_funcionario(id_funcionario):
    # Carregar o arquivo Excel
    wb = openpyxl.load_workbook(planilha_caminho)
    
    # Selecionar a planilha "Funcionarios"
    ws = wb['Funcionarios']
    
    # Percorrer as linhas da planilha a partir da segunda linha
    for row in ws.iter_rows(min_row=2):
        # Acessar o valor na primeira coluna (ID) e garantir que ambos são strings para comparação
        id_valor = str(row[0].value)
        id_funcionario_str = str(id_funcionario)
        
        # Verificar se o ID corresponde ao que queremos excluir
        if id_valor == id_funcionario_str:
            # Excluir a linha
            ws.delete_rows(row[0].row)
            break
    
    # Salvar o arquivo sem afetar outras planilhas
    wb.save(planilha_caminho)

    # Selecionar a planilha "Funcionarios"
    ws = wb['Users']
    
    # Percorrer as linhas da planilha a partir da segunda linha
    for row in ws.iter_rows(min_row=2):
        # Acessar o valor na primeira coluna (ID) e garantir que ambos são strings para comparação
        id_valor = str(row[0].value)
        id_funcionario_str = str(id_funcionario)
        
        # Verificar se o ID corresponde ao que queremos excluir
        if id_valor == id_funcionario_str:
            # Excluir a linha
            ws.delete_rows(row[0].row)
            break
    
    # Salvar o arquivo sem afetar outras planilhas
    wb.save(planilha_caminho)
    print(f"Funcionario(a) foi excluido(a) com sucesso.")

"""
Busca clientes na planilha 'Clientes' com base no nome ou ID fornecido.

Parâmetros:
- nome (str): Nome do cliente a ser buscado. Pode ser o nome completo ou o ID do cliente.

Comportamento:
- Se `nome` for diferente de 'Todos':
    - Verifica se `nome` é uma string.
    - Busca clientes cujo nome contém a string fornecida (ignora diferenças de maiúsculas e minúsculas).
    - Se nenhum cliente for encontrado pelo nome, busca também pelo ID.
    - Se encontrado, exibe os resultados; caso contrário, exibe uma mensagem informando que nenhum cliente foi encontrado.
- Se `nome` for 'Todos':
    - Retorna todos os clientes na planilha.
"""
def buscar_cliente(nome):
    df = ler_dados_planilha('Clientes')

    if nome != 'Todos':
        if not isinstance(nome, str):
            print("O nome do(a) cliente deve ser uma string.")
            return
        resultado = df[df['Nome'].str.contains(nome.strip(), case=False, na=False)]
    else:
        resultado = df

    if not resultado.empty:
        resultado = resultado.to_string(index=False)
        print("\n" + resultado + "\n")
    else:
        resultado = df[df['ID'].str.contains(nome.strip(), case=False, na=False)]
        if not resultado.empty:
            resultado = resultado.to_string(index=False)
            print("\n" + resultado + "\n")
        else:
            print("Nenhum cliente encontrado.")

"""
Atualiza as informações de um cliente na planilha 'Clientes'.

Parâmetros:
- novo_nome (str): Novo nome do cliente.
- novo_cpf (str): Novo CPF do cliente.
- novo_rg (str): Novo RG do cliente.
- nova_data_nasc (str): Nova data de nascimento do cliente.
- id_cliente (int): ID do cliente a ser atualizado.

Comportamento:
- Lê os dados da planilha 'Clientes'.
- Localiza o cliente pelo ID fornecido.
- Se encontrado, atualiza os campos fornecidos com novos valores.
- Salva as alterações na planilha.
- Imprime uma mensagem indicando se a atualização foi bem-sucedida ou se o cliente não foi encontrado.
"""
def editar_cliente(novo_nome, novo_cpf, novo_rg, nova_data_nasc, id_cliente):
    # Lê os dados atuais da planilha 'clientes'
    df_cliente = ler_dados_planilha('Clientes')

    if df_cliente is not None:
        # Localiza o cliente pelo ID
        cliente_idx = df_cliente[df_cliente['ID'] == id_cliente].index

        if not cliente_idx.empty:
            # Atualiza os campos se o usuário forneceu novos valores
            if novo_nome:
                df_cliente.at[cliente_idx[0], 'Nome'] = novo_nome
            if novo_cpf:
                df_cliente.at[cliente_idx[0], 'CPF'] = novo_cpf
            if novo_rg:
                df_cliente.at[cliente_idx[0], 'RG'] = novo_rg
            if nova_data_nasc:
                df_cliente.at[cliente_idx[0], 'Data de Nascimento'] = nova_data_nasc

            # Salva as alterações na planilha
            escrever_dados_planilha(df_cliente, 'Clientes')

            a =1
        else:
            print("cliente com o ID especificado não foi encontrado.")
    else:
        print("Erro ao ler a planilha.")

    df_cliente = ler_dados_planilha('Users')

    print("Cliente atualizado com sucesso!")

"""
Exclui um cliente da planilha 'Clientes' com base no ID fornecido.

Parâmetros:
- id_cliente (int): ID do cliente a ser excluído.

Comportamento:
- Carrega o arquivo Excel e seleciona a planilha 'Clientes'.
- Percorre as linhas da planilha e encontra a linha com o ID correspondente.
- Exclui a linha e salva o arquivo Excel.
- Imprime uma mensagem indicando o sucesso ou falha na exclusão.
"""
def excluir_cliente(id_cliente):
    # Carregar o arquivo Excel
    wb = openpyxl.load_workbook(planilha_caminho)
    
    # Selecionar a planilha "clientes"
    ws = wb['Clientes']
    
    # Percorrer as linhas da planilha a partir da segunda linha
    for row in ws.iter_rows(min_row=2):
        # Acessar o valor na primeira coluna (ID) e garantir que ambos são strings para comparação
        id_valor = str(row[0].value)
        id_cliente_str = str(id_cliente)
        
        # Verificar se o ID corresponde ao que queremos excluir
        if id_valor == id_cliente_str:
            # Excluir a linha
            ws.delete_rows(row[0].row)
            break
    
    # Salvar o arquivo sem afetar outras planilhas
    wb.save(planilha_caminho)
    print(f"Cliente foi excluido(a) com sucesso.")

"""
Processa um arquivo de texto contendo informações de uma venda e salva os dados em uma planilha Excel.

O arquivo de texto deve conter informações sobre o cliente, produtos vendidos, quantidades, subtotais,
método de pagamento e outros detalhes da venda. As informações são extraídas e organizadas em um DataFrame
do pandas, que é então salvo em um arquivo Excel.

Parâmetros:
nome_arquivo (str): O caminho para o arquivo de texto que contém as informações da venda.

Retorna:
None: Salva os dados processados em uma planilha Excel e imprime uma mensagem de sucesso.

Exceções:
- Se o arquivo não for encontrado ou houver um erro durante o processamento, a função imprime uma mensagem de erro.
"""
def processar_arquivo_venda(nome_arquivo):
    with open(nome_arquivo, 'r') as file:
        lines = file.readlines()
    
    # Inicializar variáveis
    nome_cliente = ""
    nome_vendedor = "Gabriel Tavares"  # Nome do vendedor fixo para o exemplo
    produtos_vendidos = []
    quantidades = []
    subtotal = []
    metodo_pagamento = ""
    parcelas = 1
    data_hora_venda = datetime.now().strftime("%d/%m/%Y %H:%M")
    id_compra = 1  # ID da compra fixo para o exemplo
    
    # Processar linhas do arquivo
    for i, line in enumerate(lines):
        if "Nome do(a) cliente:" in line:
            nome_cliente = line.strip().split(": ")[1]
        elif "Pagamento:" in line:
            metodo_pagamento_line = lines[i+5].strip()
            metodo_pagamento = re.findall(r'\d+', metodo_pagamento_line)[0]
        elif "CPF do(a) cliente:" in line:
            # Extrair CPF se necessário (não usado diretamente aqui)
            cpf_cliente = line.strip().split(": ")[1]
        elif line.startswith('| Produto:'):
            produto = line.split(': ')[1].strip()
            if produto:  # Ignorar produtos vazios
                produtos_vendidos.append(produto)
                quantidade_line = lines[i+1].strip()
                quantidade = float(re.findall(r'[\d.]+', quantidade_line)[0])
                quantidades.append(quantidade)
                subtotal_line = lines[i+2].strip()
                subtotal.append(float(re.findall(r'[\d.]+', subtotal_line)[0]))
    
    # Criar DataFrame
    df = pd.DataFrame({
        "ID": [id_compra] * len(produtos_vendidos),
        "Nome do Cliente": [nome_cliente] * len(produtos_vendidos),
        "Nome do Vendedor": [nome_vendedor] * len(produtos_vendidos),
        "Produtos Vendidos": produtos_vendidos,
        "Quantidade": quantidades,
        "Subtotal de cada produto": subtotal,
        "Método de Pagamento": [metodo_pagamento] * len(produtos_vendidos),
        "Quantidade de Parcelas": [parcelas] * len(produtos_vendidos),
        "Data e hora da venda": [data_hora_venda] * len(produtos_vendidos)
    })

    df.to_excel(planilha_caminho, index=False, engine='openpyxl')

"""
Executa funções baseadas nos argumentos passados pela linha de comando.

Dependendo dos argumentos fornecidos, o script executa ações como excluir, salvar, buscar, editar, cadastrar,
e validar login. Cada comando é associado a uma função específica que manipula dados relacionados a produtos,
funcionários e clientes.

Exceções:
- Se a quantidade de argumentos não corresponder a nenhum dos casos esperados, o script imprime uma mensagem de uso
    e encerra com código de status 1.
"""
if __name__ == "__main__":
    if "excluir" in sys.argv and "produto" in sys.argv:
        id_produto = sys.argv[3]
        excluir_produto(id_produto)
    if "salvar" in sys.argv and "compra" in sys.argv:
        nome_arquivo = sys.argv[3]
        processar_arquivo_venda(nome_arquivo)
    elif "excluir" in sys.argv and "funcionario" in sys.argv:
        id_funcionario = sys.argv[3]
    elif "excluir" in sys.argv and "cliente" in sys.argv:
        id_cliente = sys.argv[3]
        excluir_cliente(id_cliente)
    elif "buscar" in sys.argv and "funcionario" in sys.argv:
        funcionario = sys.argv[3]
        buscar_funcionario(funcionario)
    elif "buscar" in sys.argv and "cliente" in sys.argv:
        cliente = sys.argv[3]
        buscar_cliente(cliente)
    elif "editar" in sys.argv and "funcionario" in sys.argv:
        novo_nome = sys.argv[4]
        novo_cpf = sys.argv[5]
        novo_rg = sys.argv[6]
        nova_data_nasc = sys.argv[7]
        novo_user = sys.argv[8]
        nova_senha = sys.argv[9]
        id_funcionario = sys.argv[3]
        editar_funcionario(novo_nome, novo_cpf, novo_rg, nova_data_nasc, novo_user, nova_senha, id_funcionario)
    elif "editar" in sys.argv and "cliente" in sys.argv:
        novo_nome = sys.argv[4]
        novo_cpf = sys.argv[5]
        novo_rg = sys.argv[6]
        nova_data_nasc = sys.argv[7]
        id_cliente = sys.argv[3]
        editar_cliente(novo_nome, novo_cpf, novo_rg, nova_data_nasc, id_cliente)
    elif "editar" in sys.argv and "produto" in sys.argv:
        novo_nome = sys.argv[2]
        novo_preco = sys.argv[3]
        nova_quantidade = sys.argv[4]
        tipo_venda = sys.argv[5]
        id_produto = sys.argv[1]
        editar_produto(novo_nome, novo_preco, nova_quantidade, tipo_venda, id_produto)
    elif "cadastrar" in sys.argv and "cliente" in sys.argv:
        nome = sys.argv[3]
        cpf = str(sys.argv[4])  # Certifica-se de que o CPF seja tratado como string
        data_nascimento = sys.argv[5]
        rg = sys.argv[6]
        cadastrar_cliente(nome, cpf, data_nascimento, rg)
    elif len(sys.argv) == 8:
        nome = sys.argv[1]
        cpf = str(sys.argv[2])  # Certifica-se de que o CPF seja tratado como string
        data_nascimento = sys.argv[3]
        rg = sys.argv[4]
        username = sys.argv[5]
        password = sys.argv[6]
        tipo = sys.argv[7]
        cadastrar_funcionario(nome, cpf, data_nascimento, rg, username, password, tipo)
    elif len(sys.argv) == 3:
        username = sys.argv[1]
        password = sys.argv[2]
        if validar_login(username, password):
            sys.exit(0)  # Login bem-sucedido
        else:
            sys.exit(1)  # Falha no login
    elif len(sys.argv) == 5:
        nome = sys.argv[1] # Certifica-se de que o CPF seja tratado como string
        qtd_produto = sys.argv[2]
        preco = str(sys.argv[3])
        tipo_produto = str(sys.argv[4])
        cadastrar_produto(nome, preco, qtd_produto, tipo_produto)
    elif len(sys.argv) == 2:
        nome = sys.argv[1]
        buscar_produto(nome)
    else:
        print("Uso: python excel_utils.py <username> <password> ou python excel_utils.py <nome> <cpf> <data_nascimento> <rg> <username> <password> <tipo>")
        print(len(sys.argv))
        print(sys.argv)
        sys.exit(1)